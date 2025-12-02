import os
import fitz  # PyMuPDF
from rapidocr_onnxruntime import RapidOCR
import docx
import openpyxl
from tqdm import tqdm  # Barra de progreso

class FileProcessor:
    def __init__(self):
        # Iniciamos el motor OCR una sola vez
        self.ocr = RapidOCR()

    def process_file(self, file_path):
        ext = os.path.splitext(file_path)[1].lower()
        
        if ext == ".pdf":
            return self._process_pdf(file_path)
        elif ext in [".docx", ".doc"]:
            return self._process_word(file_path)
        elif ext in [".xlsx", ".xls"]:
            return self._process_excel(file_path)
        elif ext in [".jpg", ".png", ".jpeg"]:
            return self._process_image(file_path)
        elif ext in [".txt", ".md"]:
            with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                return f.read()
        return ""

    def _process_pdf(self, path):
        text_content = ""
        doc = fitz.open(path)
        total_pages = len(doc)
        
        # BARRA DE PROGRESO POR PÁGINA
        print(f"      📄 Analizando PDF ({total_pages} pág)...")
        for page in tqdm(doc, total=total_pages, unit="pág", leave=False):
            text = page.get_text()
            
            # Si hay poco texto, aplicamos OCR (GPU)
            if len(text.strip()) < 50:
                pix = page.get_pixmap(matrix=fitz.Matrix(1.5, 1.5))
                img_bytes = pix.tobytes("png")
                ocr_result, _ = self.ocr(img_bytes)
                if ocr_result:
                    for line in ocr_result:
                        text += line[1] + "\n"
            
            text_content += text + "\n"
        return text_content

    def _process_image(self, path):
        result, _ = self.ocr(path)
        text = ""
        if result:
            for line in result:
                text += line[1] + "\n"
        return text

    def _process_word(self, path):
        doc = docx.Document(path)
        return "\n".join([p.text for p in doc.paragraphs if p.text.strip()])

    def _process_excel(self, path):
        wb = openpyxl.load_workbook(path, data_only=True)
        text = ""
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(values_only=True):
                row_text = " | ".join([str(c) for c in row if c])
                if row_text:
                    text += row_text + "\n"
        return text